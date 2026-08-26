"""订单↔门店监控测试：提交订单→登记共享表（不再自动延期）；状态变化→群通知。

v4.1（2026-08-14 用户决策）：下单不延期，等货照常算时效；状态含「签收」→
标记到货、开始计时，此后每日提醒直至完成。
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pytest

from db import Database
from models import NormalizedMessage
from notifier import Notifier
from pipeline import MessageProcessingPipeline, RuntimeMode
from routing.pending_actions import PendingActionService
from routing.ticket_contexts import TicketContextStore
from routing.ticket_router import TicketRouter
from semantics.protocol_loader import load_protocol
from tickets.executor import TicketCommandExecutor
from tickets.repository import TicketRepository
from workers.scheduler import SchedulerWorker

_PROTOCOL_PATH = Path(__file__).resolve().parent.parent / "protocols" / "ticket_semantics.v4.json"
GROUP = {"group_id": "G1", "store_name": "测试店",
         "manager_ids": ["mgr"], "engineer_ids": ["eng"], "other_member_ids": ["staff"]}


class _OrderFakeClassifier:
    """模拟全AI对 # 语法及裸订单号的识别，供订单流测试使用。"""
    def __init__(self, protocol):
        self.protocol = protocol

    async def classify(self, message, candidates=None, pending_action=None, history=None):
        from semantics.types import SemanticDecision
        text = (message.content or "").strip()
        if text.startswith("#") and self.protocol is not None:
            try:
                from semantics.keyword_matcher import match_keyword as _mk
                kw = _mk(text, self.protocol)
                if kw is not None:
                    return SemanticDecision(
                        protocol_version=kw.protocol_version, source="SEMANTIC_MODEL",
                        intent=kw.intent, target_ticket_no=kw.target_ticket_no,
                        intent_confidence=0.95, fields=dict(kw.fields),
                        missing_fields=kw.missing_fields, evidence=kw.evidence,
                    )
            except Exception:
                pass
        # 裸订单号 / 诊断+订单 混合
        import re as _re
        order_tokens = _re.findall(r"(?<![A-Za-z0-9-])[A-Za-z0-9-]{6,64}(?![A-Za-z0-9-])", text or "")
        orders = [
            t for t in order_tokens
            if sum(ch.isdigit() for ch in t) >= 6
            and not _re.fullmatch(r"1[3-9]\d{9}", t)
        ]
        if orders:
            # 只要含订单号且候选中或消息明显像订单提交，就视为 repair_plan.submit
            if any(kw in text for kw in ("订单", "单号", "采购", "TB-", "估计", "铰链", "坏")) or len(orders) >= 1:
                # 提取诊断（若有）
                fields: dict = {"order_no": orders[0], "order_nos": orders}
                # 简易诊断提取：含铰链/故障等则加入
                if any(cue in text for cue in ("铰链", "判断", "估计", "应该", "可能", "坏", "故障")):
                    # 取整句作为诊断
                    diag = text
                    for o in orders:
                        diag = diag.replace(o, " ")
                    diag = diag.strip()
                    if diag:
                        fields["diagnosis_items"] = [diag[:100]]
                return SemanticDecision(
                    protocol_version="4.0.0", source="SEMANTIC_MODEL",
                    intent="ticket.repair_plan.submit", target_ticket_no=None,
                    intent_confidence=0.9, fields=fields,
                )
        return SemanticDecision(protocol_version="4.0.0", source="SEMANTIC_MODEL",
                                 intent="chat.ignore", target_ticket_no=None, intent_confidence=0.0)


@pytest.fixture()
def env(tmp_path, monkeypatch):
    db = Database(tmp_path / "order.db")
    db.init_schema()
    db.upsert_group(GROUP)
    shared = tmp_path / "订单门店状态表.xlsx"
    monkeypatch.setattr("config.ORDER_STORE_TABLE_PATH", shared)
    protocol = load_protocol(_PROTOCOL_PATH)
    repo = TicketRepository(db)
    router = TicketRouter()
    context = TicketContextStore(db)
    pending = PendingActionService(db)
    executor = TicketCommandExecutor(db, repo)
    sent: list[str] = []
    notifier = Notifier(db, lambda target, text: sent.append(text))
    classifier = _OrderFakeClassifier(protocol)
    pipeline = MessageProcessingPipeline(
        db=db, repo=repo, protocol=protocol, router=router, context=context,
        pending=pending, executor=executor, notifier=notifier,
        classifier=classifier, mode=RuntimeMode.PRODUCTION,
    )
    worker = SchedulerWorker(db=db, notifier=notifier, interval=60)
    yield SimpleNamespace(db=db, sent=sent, pipeline=pipeline, worker=worker, shared=shared)
    db.close()


class SimpleNamespace:
    def __init__(self, **kw):
        self.__dict__.update(kw)


async def _create_ticket(env) -> dict:
    msg = NormalizedMessage(message_id="c1", group_id="G1", sender_id="staff", sender_name="店员",
                            content="#报修\n主题：金库\n位置：五房\n问题描述：风机风力小\n时效：3天",
                            message_type="text", sent_at=datetime.now(), sender_role="OTHER")
    env.db.enqueue_message(msg)
    row = env.db.connect().execute("SELECT * FROM inbox_messages WHERE message_id='c1'").fetchone()
    await env.pipeline.process(dict(row))
    return env.db.list_active_tickets("G1")[0]


async def _submit_order(env, order_no: str, message_id: str = "r1"):
    msg = NormalizedMessage(message_id=message_id, group_id="G1", sender_id="eng", sender_name="工程师",
                            content=f"#维修方式\n维修方式：淘宝采购后自行维修\n订单号：{order_no}",
                            message_type="text", sent_at=datetime.now(), sender_role="ENGINEER")
    env.db.enqueue_message(msg)
    row = env.db.connect().execute("SELECT * FROM inbox_messages WHERE message_id=?", (message_id,)).fetchone()
    await env.pipeline.process(dict(row))


def _set_shared_status(env, order_no: str, status: str, tracking: str = "") -> None:
    wb = openpyxl.load_workbook(env.shared)
    ws = wb.active
    for r in ws.iter_rows(min_row=2):
        if r[0].value == order_no:
            r[3].value = status
            r[4].value = tracking
    wb.save(env.shared)
    wb.close()


@pytest.mark.asyncio
async def test_order_submit_registers_without_extension(env):
    """v4.1：提交订单只登记，不再自动延期 3 天（等货照常算时效）。"""
    ticket = await _create_ticket(env)
    before = ticket["current_deadline_at"]
    await _submit_order(env, "TB-2024-0001")
    fresh = env.db.get_ticket(ticket["id"])
    assert fresh["current_deadline_at"] == before  # 不再延期
    monitor = env.db.get_order_monitor("TB-2024-0001")
    assert monitor is not None and monitor["ticket_no"] == ticket["ticket_no"]
    # 共享表有该订单行
    wb = openpyxl.load_workbook(env.shared)
    ws = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == "TB-2024-0001"]
    wb.close()
    assert len(rows) == 1
    assert rows[0][1] == "测试店" and rows[0][2] == ticket["ticket_no"]
    # 2026-08-24 静默化：订单登记成功属纯告知，不再群内回执
    assert not any("已登记" in s for s in env.sent)
    assert not any("自动延期" in s for s in env.sent)


@pytest.mark.asyncio
async def test_order_submit_by_any_role(env):
    """订单号人人可发（店长/工程师/其他成员），不只工程师（2026-08-12）。"""
    await _create_ticket(env)
    msg = NormalizedMessage(message_id="r-staff", group_id="G1", sender_id="staff", sender_name="店员",
                            content="#维修方式\n维修方式：淘宝采购后自行维修\n订单号：TB-ANY-0001",
                            message_type="text", sent_at=datetime.now(), sender_role="OTHER")
    env.db.enqueue_message(msg)
    row = env.db.connect().execute("SELECT * FROM inbox_messages WHERE message_id='r-staff'").fetchone()
    await env.pipeline.process(dict(row))
    assert env.db.get_order_monitor("TB-ANY-0001") is not None
    # 静默化：登记成功不回执
    assert not any("已登记" in s for s in env.sent)


@pytest.mark.asyncio
async def test_bare_order_number_registers_without_extension(env):
    """群里只发一个订单号 → 视为提交订单：登记但不延期（2026-08-14）。"""
    ticket = await _create_ticket(env)
    before = ticket["current_deadline_at"]
    msg = NormalizedMessage(message_id="r-bare", group_id="G1", sender_id="eng", sender_name="工程师",
                            content="单号 5125938806116169335",
                            message_type="text", sent_at=datetime.now(), sender_role="ENGINEER")
    env.db.enqueue_message(msg)
    row = env.db.connect().execute("SELECT * FROM inbox_messages WHERE message_id='r-bare'").fetchone()
    await env.pipeline.process(dict(row))
    # 订单登记，工单不延期
    monitor = env.db.get_order_monitor("5125938806116169335")
    assert monitor is not None and monitor["ticket_id"] == ticket["id"]
    fresh = env.db.get_ticket(ticket["id"])
    assert fresh["current_deadline_at"] == before
    # 不写空的维修方式版本
    rows = env.db.connect().execute(
        "SELECT COUNT(*) FROM repair_method_versions WHERE ticket_id=?", (ticket["id"],)
    ).fetchone()[0]
    assert rows == 0
    # 静默化：裸单号登记成功不回执，也不弹「已记录维修方式」
    assert not any("已登记" in s for s in env.sent)
    assert not any("已记录维修方式" in s for s in env.sent)


@pytest.mark.asyncio
async def test_multiple_orders_with_diagnosis_registered(env):
    """「估计是铰链坏了，采购了2个，单号是X和Y」→ 全部订单登记 + 记录诊断（不再延期）。"""
    ticket = await _create_ticket(env)
    before = ticket["current_deadline_at"]
    msg = NormalizedMessage(
        message_id="r-multi", group_id="G1", sender_id="eng", sender_name="工程师",
        content="估计是铰链坏了，采购了2个，单号是5127629004214178517和5127628896203022943",
        message_type="text", sent_at=datetime.now(), sender_role="ENGINEER",
    )
    env.db.enqueue_message(msg)
    row = env.db.connect().execute("SELECT * FROM inbox_messages WHERE message_id='r-multi'").fetchone()
    await env.pipeline.process(dict(row))
    # 两个订单都登记
    assert env.db.get_order_monitor("5127629004214178517") is not None
    assert env.db.get_order_monitor("5127628896203022943") is not None
    # 工单不延期
    fresh = env.db.get_ticket(ticket["id"])
    assert fresh["current_deadline_at"] == before
    # 诊断一并记录
    diag = env.db.connect().execute(
        "SELECT items_json FROM diagnosis_versions WHERE ticket_id=? AND is_current=1", (ticket["id"],)
    ).fetchone()
    assert diag is not None and "铰链" in diag["items_json"]
    # 静默化：多订单登记成功也不回执
    assert not any(
        "5127629004214178517" in s and "5127628896203022943" in s for s in env.sent
    )


@pytest.mark.asyncio
async def test_duplicate_order_registered_once(env):
    ticket = await _create_ticket(env)
    before = ticket["current_deadline_at"]
    await _submit_order(env, "TB-2024-0001", "r1")
    after_first = env.db.get_ticket(ticket["id"])["current_deadline_at"]
    await _submit_order(env, "TB-2024-0001", "r2")  # 重复提交同单号
    after_second = env.db.get_ticket(ticket["id"])["current_deadline_at"]
    assert after_first == after_second == before  # 不延期、不重复登记
    count = env.db.connect().execute(
        "SELECT COUNT(*) FROM order_monitor WHERE order_id='TB-2024-0001'").fetchone()[0]
    assert count == 1


@pytest.mark.asyncio
async def test_scan_notifies_shipped_once(env):
    await _create_ticket(env)
    await _submit_order(env, "TB-2024-0001")
    _set_shared_status(env, "TB-2024-0001", "卖家已发货", "SF-1")
    env.worker.scan_order_status()
    assert any("已发货" in s and "TB-2024-0001" in s for s in env.sent)
    env.worker.scan_order_status()  # 再扫一次，不重复通知
    shipped_count = sum(1 for s in env.sent if "已发货" in s)
    assert shipped_count == 1


@pytest.mark.asyncio
async def test_scan_notifies_closed_unpaid(env):
    await _create_ticket(env)
    await _submit_order(env, "TB-2024-0002")
    _set_shared_status(env, "TB-2024-0002", "交易关闭")
    env.worker.scan_order_status()
    assert any("因未付款已关闭" in s and "TB-2024-0002" in s for s in env.sent)
    # 第二次扫描不重复
    env.worker.scan_order_status()
    assert sum(1 for s in env.sent if "未付款已关闭" in s) == 1


@pytest.mark.asyncio
async def test_scan_ignores_status_without_change(env):
    await _create_ticket(env)
    await _submit_order(env, "TB-2024-0003")
    _set_shared_status(env, "TB-2024-0003", "等待买家付款")
    env.worker.scan_order_status()
    _set_shared_status(env, "TB-2024-0003", "等待买家付款")  # 状态未变
    env.worker.scan_order_status()
    # 「等待买家付款」不触发发货/关闭通知
    assert not any("已发货" in s for s in env.sent)
    assert not any("关闭" in s for s in env.sent)
    monitor = env.db.get_order_monitor("TB-2024-0003")
    assert monitor["last_status"] == "等待买家付款"


@pytest.mark.asyncio
async def test_scan_received_marks_order_and_notifies_once(env):
    """状态含「签收」→ 标记到货、提醒一次「开始计时」，不重复。"""
    await _create_ticket(env)
    await _submit_order(env, "TB-2024-0004")
    _set_shared_status(env, "TB-2024-0004", "已签收")
    env.worker.scan_order_status()
    assert any("已签收到货" in s and "开始计时" in s for s in env.sent)
    monitor = env.db.get_order_monitor("TB-2024-0004")
    assert monitor["received_at"] is not None
    assert monitor["received_notified"] == 1
    env.worker.scan_order_status()  # 再扫一次不重复
    assert sum(1 for s in env.sent if "已签收到货" in s) == 1


@pytest.mark.asyncio
async def test_scan_received_via_trade_success(env):
    """「交易成功」（不带签收二字）同样触发到货计时。"""
    await _create_ticket(env)
    await _submit_order(env, "TB-2024-0007")
    _set_shared_status(env, "TB-2024-0007", "交易成功")
    env.worker.scan_order_status()
    assert any("已签收到货" in s for s in env.sent)
    assert env.db.get_order_monitor("TB-2024-0007")["received_at"] is not None


@pytest.mark.asyncio
async def test_received_daily_reminder_once_per_day(env):
    """签收后每天提醒一次，同日去重，次日再提醒。"""
    await _create_ticket(env)
    await _submit_order(env, "TB-2024-0005")
    _set_shared_status(env, "TB-2024-0005", "已签收")
    env.worker.scan_order_status()
    env.sent.clear()

    day1 = datetime.now()
    env.worker.scan_received_reminders(day1)
    assert any("已签收" in s and "仍在处理" in s for s in env.sent)
    day1_count = sum(1 for s in env.sent if "仍在处理" in s)
    assert day1_count == 1

    env.worker.scan_received_reminders(day1)  # 同日再扫不重复
    assert sum(1 for s in env.sent if "仍在处理" in s) == day1_count

    env.worker.scan_received_reminders(day1 + timedelta(days=1))  # 次日再提醒
    assert sum(1 for s in env.sent if "仍在处理" in s) == day1_count + 1


@pytest.mark.asyncio
async def test_sla_reminder_skips_received_ticket(env):
    """签收后的工单不再按 SLA 截止提醒（改走每日提醒）。"""
    ticket = await _create_ticket(env)
    await _submit_order(env, "TB-2024-0006")
    _set_shared_status(env, "TB-2024-0006", "已签收")
    env.worker.scan_order_status()
    env.sent.clear()
    # 把截止时间改到过去，模拟已超时
    env.db.connect().execute(
        "UPDATE tickets SET current_deadline_at='2026-01-01 00:00:00' WHERE id=?", (ticket["id"],)
    )
    env.worker.scan_sla_reminders(datetime.now())
    assert not any("超时效" in s for s in env.sent)
    assert not any("时效即将到期" in s for s in env.sent)


# ─────────────────── 共享表写入兜底（2026-08-25） ───────────────────
@pytest.mark.asyncio
async def test_order_submit_marks_xlsx_synced(env):
    """登记成功（或行已存在）后，订单标记 xlsx_synced=1。"""
    await _create_ticket(env)
    await _submit_order(env, "TB-SYNC-0001")
    monitor = env.db.get_order_monitor("TB-SYNC-0001")
    assert monitor is not None
    assert monitor["xlsx_synced"] == 1


@pytest.mark.asyncio
async def test_order_submit_tolerates_shared_table_failure(env, monkeypatch):
    """共享表被占用写失败 → 不阻断登记、不向群里抛错，订单留在待补同步队列。"""
    await _create_ticket(env)

    def _boom(*_args, **_kw):
        raise PermissionError(1, "Operation not permitted")

    monkeypatch.setattr("reconciling.order_store.append_order_row", _boom)
    await _submit_order(env, "TB-LOCK-0001")  # 不应抛异常
    monitor = env.db.get_order_monitor("TB-LOCK-0001")
    assert monitor is not None and monitor["xlsx_synced"] == 0


@pytest.mark.asyncio
async def test_scheduler_resyncs_failed_shared_rows(env, monkeypatch):
    """故障解除后调度器补同步：把未同步订单补进共享表并翻转标记；无待同步时不动。"""
    import reconciling.order_store as order_store

    await _create_ticket(env)
    real = order_store.append_order_row
    state = {"fail": True}

    def _flaky(*args, **kwargs):
        if state["fail"]:
            raise PermissionError(1, "Operation not permitted")
        return real(*args, **kwargs)

    monkeypatch.setattr(order_store, "append_order_row", _flaky)
    await _submit_order(env, "TB-RETRY-0001")
    assert env.db.get_order_monitor("TB-RETRY-0001")["xlsx_synced"] == 0

    # 故障解除 → 一轮调度扫描即补写成功
    state["fail"] = False
    fixed = env.worker.scan_shared_table_resync(datetime.now())
    assert fixed == 1
    wb = openpyxl.load_workbook(env.shared)
    ws = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == "TB-RETRY-0001"]
    wb.close()
    assert len(rows) == 1 and rows[0][2] is not None
    assert env.db.get_order_monitor("TB-RETRY-0001")["xlsx_synced"] == 1

    # 已全部同步后再扫：无事可做
    assert env.worker.scan_shared_table_resync(datetime.now()) == 0
