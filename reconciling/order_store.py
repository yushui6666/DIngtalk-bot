"""订单↔门店共享表（xlsx）读写。

与另一个 AI 的协作契约：
- 表文件：``config.ORDER_STORE_TABLE_PATH``（默认 /Desktop/淘宝对账/订单门店状态表.xlsx）
- 列：``order_id | store | ticket_no | status | tracking_number | updated_at``
- 我们（报修系统）写入前三列（工单提交订单号时追加）；
- 另一个 AI 每天拉淘宝状态，只更新 ``status`` / ``tracking_number`` / ``updated_at``；
- 双方写入前都必须先读最新文件再改，避免覆盖对方更新。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from logger import get_logger

logger = get_logger(__name__)

_HEADERS = ("order_id", "store", "ticket_no", "status", "tracking_number", "updated_at")


def _ensure_file(path: Path) -> None:
    import openpyxl

    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "orders"
    ws.append(list(_HEADERS))
    wb.save(path)


def append_order_row(path: Path, *, order_id: str, store: str, ticket_no: str) -> bool:
    """追加一行（order 已存在则跳过，幂等）。"""
    import openpyxl

    _ensure_file(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    existing = {str(r[0]) for r in ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}
    if order_id in existing:
        wb.close()
        return False
    ws.append([order_id, store, ticket_no, "", "", ""])
    wb.save(path)
    wb.close()
    logger.info("订单已写入共享表 order=%s store=%s ticket=%s", order_id, store, ticket_no)
    return True


def read_order_rows(path: Path) -> list[dict[str, Any]]:
    """读取共享表所有行（跳过表头）。文件不存在返回空列表。"""
    import openpyxl

    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        data.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return data
